#!/usr/bin/env python3
"""
==============================================================================
MedVisionSort - Appium E2E Live Automation Testing Suite
==============================================================================
Application: MedVisionSort (Android App)
Package:     com.example.medicalsorter
Activity:    com.example.medicalsorter.MainActivity
Credentials: saagar@gmail.com / admin123
Report:      Excel (.xlsx) with styled PASS/FAIL results per test case
==============================================================================
"""

import time
import datetime
import traceback
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from contextlib import contextmanager

from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
from appium.options.android import UiAutomator2Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# ==============================================================================
# Configuration
# ==============================================================================
APK_PATH = "/Users/saagar/Projects/pdd_v2/mobile_app/app/build/outputs/apk/debug/app-debug.apk"
SCREENSHOTS_DIR = "/Users/saagar/Projects/pdd_v2/screenshots"
REPORT_PATH = "/Users/saagar/Projects/pdd_v2/E2E_Test_Report_MedVisionSort_Appium.xlsx"

WAIT_TIMEOUT = 10
SHORT_WAIT = 4

# Create directories if they do not exist
os.makedirs(SCREENSHOTS_DIR, exist_ok=True)

# ==============================================================================
# Test Result Container
# ==============================================================================
class TestResult:
    def __init__(self, tc_id, module, test_name, description, steps, expected):
        self.tc_id       = tc_id
        self.module      = module
        self.test_name   = test_name
        self.description = description
        self.steps       = steps
        self.expected    = expected
        self.status      = "NOT RUN"
        self.actual      = ""
        self.error       = ""
        self.duration    = 0.0
        self.timestamp   = ""

results: list[TestResult] = []
driver = None

# ==============================================================================
# Helpers
# ==============================================================================
def setup_driver():
    global driver
    print("[DRIVER] Connecting to Appium Server...")
    options = UiAutomator2Options()
    options.platform_name = 'Android'
    options.device_name = 'emulator-5554'
    options.app = APK_PATH
    options.app_package = 'com.example.medicalsorter'
    options.app_activity = 'com.example.medicalsorter.MainActivity'
    options.automation_name = 'UiAutomator2'
    options.auto_grant_permissions = True
    options.no_reset = False
    options.new_command_timeout = 300
    
    driver = webdriver.Remote('http://localhost:4723', options=options)
    print("[DRIVER] Appium Android WebDriver initialised ✓")
    return driver

def wait_for(locator, by=AppiumBy.XPATH, timeout=WAIT_TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, locator))
    )

def wait_visible(locator, by=AppiumBy.XPATH, timeout=WAIT_TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((by, locator))
    )

def wait_clickable(locator, by=AppiumBy.XPATH, timeout=WAIT_TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, locator))
    )

def find_by_text(text, timeout=SHORT_WAIT):
    return wait_visible(f"//*[@text='{text}']", AppiumBy.XPATH, timeout)

def find_contains_text(text, timeout=SHORT_WAIT):
    return wait_visible(f"//*[contains(@text, '{text}')]", AppiumBy.XPATH, timeout)

def click_by_text(text, timeout=SHORT_WAIT):
    el = wait_clickable(f"//*[@text='{text}']", AppiumBy.XPATH, timeout)
    el.click()
    return el

def click_by_accessibility_id(acc_id, timeout=SHORT_WAIT):
    el = wait_clickable(acc_id, AppiumBy.ACCESSIBILITY_ID, timeout)
    el.click()
    return el

def get_edit_texts():
    return driver.find_elements(by=AppiumBy.CLASS_NAME, value="android.widget.EditText")

def hide_keyboard():
    try:
        if driver.is_keyboard_shown():
            driver.back()
            time.sleep(0.5)
    except Exception:
        try:
            driver.hide_keyboard()
            time.sleep(0.5)
        except Exception:
            try:
                driver.press_keycode(4) # BACK key
                time.sleep(0.5)
            except Exception:
                pass

def clear_and_type(element, text):
    element.click()
    element.clear()
    current_text = element.text
    if current_text and current_text != text:
        # Clear using keycodes if clear() didn't work
        for _ in range(len(current_text) + 3):
            driver.press_keycode(67) # KEYCODE_DEL
    element.send_keys(text)
    hide_keyboard()

# ==============================================================================
# Context Manager for Test Cases
# ==============================================================================
@contextmanager
def test_case(tc_id, module, test_name, description, steps, expected):
    tr = TestResult(tc_id, module, test_name, description, steps, expected)
    tr.timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    start = time.time()
    print(f"\n[TEST] {tc_id}: {test_name}")
    try:
        yield tr
        tr.status = "PASS"
        if not tr.actual:
            tr.actual = "Behaved as expected"
        print(f"       ✅ PASS  ({round(time.time()-start,2)}s)")
    except AssertionError as e:
        tr.status = "FAIL"
        tr.actual = str(e) or "Assertion failed"
        tr.error  = traceback.format_exc(limit=3)
        print(f"       ❌ FAIL  ({round(time.time()-start,2)}s): {tr.actual}")
        take_failed_screenshot(tc_id)
    except (TimeoutException, NoSuchElementException) as e:
        tr.status = "FAIL"
        tr.actual = f"{e.__class__.__name__}: element not found / timed out"
        tr.error  = str(e)[:300]
        print(f"       ❌ FAIL  ({round(time.time()-start,2)}s): {tr.actual}")
        take_failed_screenshot(tc_id)
    except Exception as e:
        tr.status = "FAIL"
        tr.actual = str(e)[:200]
        tr.error  = traceback.format_exc(limit=5)
        print(f"       ❌ FAIL  ({round(time.time()-start,2)}s): {tr.actual}")
        take_failed_screenshot(tc_id)
    finally:
        tr.duration = round(time.time() - start, 2)
        results.append(tr)

def take_failed_screenshot(tc_id):
    try:
        screenshot_path = os.path.join(SCREENSHOTS_DIR, f"{tc_id.lower()}_failed.png")
        driver.save_screenshot(screenshot_path)
        print(f"       📸 Saved error screenshot to: {screenshot_path}")
    except Exception as e:
        print(f"       ⚠️ Failed to save screenshot: {e}")

# ==============================================================================
# Excel Report Generator
# ==============================================================================
def generate_excel_report():
    wb = openpyxl.Workbook()

    # Colors
    C_PASS        = "FF22C55E"   # green
    C_FAIL        = "FFEF4444"   # red
    C_NOT_RUN     = "FFAAAAAA"   # grey
    C_HEADER_BG   = "FF1E293B"   # dark navy
    C_HEADER_FG   = "FFFFFFFF"   # white
    C_TITLE_BG    = "FF0F172A"   # darker navy
    C_ALT_ROW     = "FFF1F5F9"   # light blue-grey
    C_SECTION_BG  = "FF334155"   # section header
    C_ACCENT      = "FF38BDF8"   # sky blue accent

    thin = Side(style="thin", color="FFCCCCCC")
    medium = Side(style="medium", color="FF64748B")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_header = Border(left=medium, right=medium, top=medium, bottom=medium)

    def make_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def hdr_font(sz=11, bold=True, color="FFFFFFFF"):
        return Font(name="Calibri", size=sz, bold=bold, color=color)

    def body_font(sz=10, bold=False, color="FF1E293B"):
        return Font(name="Calibri", size=sz, bold=bold, color=color)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # --------------------------------------------------------------------------
    # SHEET 1 — SUMMARY
    # --------------------------------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "📊 Summary"

    ws_sum.merge_cells("A1:H1")
    ws_sum["A1"] = "🏥  MedVisionSort — Appium Mobile E2E Test Report"
    ws_sum["A1"].font      = Font(name="Calibri", size=18, bold=True, color="FFFFFFFF")
    ws_sum["A1"].fill      = make_fill(C_TITLE_BG)
    ws_sum["A1"].alignment = center
    ws_sum.row_dimensions[1].height = 42

    ws_sum.merge_cells("A2:H2")
    ws_sum["A2"] = f"Generated: {datetime.datetime.now().strftime('%d %B %Y  %H:%M:%S')}   |   Scope: Android Mobile E2E   |   Target Device: emulator-5554"
    ws_sum["A2"].font      = Font(name="Calibri", size=10, color="FFcbd5e1")
    ws_sum["A2"].fill      = make_fill("FF1E293B")
    ws_sum["A2"].alignment = center
    ws_sum.row_dimensions[2].height = 22

    total   = len(results)
    passed  = sum(1 for r in results if r.status == "PASS")
    failed  = sum(1 for r in results if r.status == "FAIL")
    not_run = sum(1 for r in results if r.status == "NOT RUN")
    rate    = (passed / total * 100) if total else 0

    stats_data = [
        ("Total Tests",    total,           "FF334155"),
        ("✅ Passed",       passed,          "FF166534"),
        ("❌ Failed",       failed,          "FF991B1B"),
        ("⏭ Not Run",      not_run,         "FF374151"),
        ("Pass Rate",       f"{rate:.1f}%",  "FF075985"),
    ]

    for col_i, (label, val, color) in enumerate(stats_data, start=2):
        col = get_column_letter(col_i)
        ws_sum.merge_cells(f"{col}4:{col}5")
        ws_sum[f"{col}4"] = label
        ws_sum[f"{col}4"].font      = Font(name="Calibri", size=9, bold=True, color="FFCBD5E1")
        ws_sum[f"{col}4"].fill      = make_fill(color)
        ws_sum[f"{col}4"].alignment = center

        ws_sum.merge_cells(f"{col}6:{col}7")
        ws_sum[f"{col}6"] = val
        ws_sum[f"{col}6"].font      = Font(name="Calibri", size=22, bold=True, color="FFFFFFFF")
        ws_sum[f"{col}6"].fill      = make_fill(color)
        ws_sum[f"{col}6"].alignment = center
        ws_sum.column_dimensions[col].width = 18

    ws_sum.row_dimensions[4].height = 20
    ws_sum.row_dimensions[5].height = 20
    ws_sum.row_dimensions[6].height = 35
    ws_sum.row_dimensions[7].height = 35

    modules = {}
    for r in results:
        m = r.module
        modules.setdefault(m, {"pass": 0, "fail": 0, "total": 0})
        modules[m]["total"] += 1
        if r.status == "PASS":
            modules[m]["pass"] += 1
        else:
            modules[m]["fail"] += 1

    row = 9
    headers = ["Module", "Total", "Passed", "Failed", "Pass Rate", "Status"]
    for ci, h in enumerate(headers, start=2):
        c = get_column_letter(ci)
        ws_sum[f"{c}{row}"] = h
        ws_sum[f"{c}{row}"].font      = hdr_font(10)
        ws_sum[f"{c}{row}"].fill      = make_fill(C_HEADER_BG)
        ws_sum[f"{c}{row}"].alignment = center
        ws_sum[f"{c}{row}"].border    = border_header
    ws_sum.row_dimensions[row].height = 22

    for ri, (mod, s) in enumerate(modules.items()):
        row += 1
        pr = (s["pass"] / s["total"] * 100) if s["total"] else 0
        status_txt = "✅ PASS" if s["fail"] == 0 else "❌ FAIL"
        row_vals = [mod, s["total"], s["pass"], s["fail"], f"{pr:.0f}%", status_txt]
        fill = make_fill(C_ALT_ROW if ri % 2 == 0 else "FFFFFFFF")
        for ci, val in enumerate(row_vals, start=2):
            c = get_column_letter(ci)
            ws_sum[f"{c}{row}"] = val
            ws_sum[f"{c}{row}"].font      = body_font()
            ws_sum[f"{c}{row}"].fill      = fill
            ws_sum[f"{c}{row}"].alignment = center
            ws_sum[f"{c}{row}"].border    = border_all
            if ci == 7:
                color = C_PASS if "PASS" in str(val) else C_FAIL
                ws_sum[f"{c}{row}"].fill = make_fill(color)
                ws_sum[f"{c}{row}"].font = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
        ws_sum.row_dimensions[row].height = 20

    # --------------------------------------------------------------------------
    # SHEET 2 — DETAILED RESULTS
    # --------------------------------------------------------------------------
    ws = wb.create_sheet("🧪 Test Cases")

    ws.merge_cells("A1:K1")
    ws["A1"] = "🧪  Detailed Mobile Test Case Results — MedVisionSort"
    ws["A1"].font      = Font(name="Calibri", size=14, bold=True, color="FFFFFFFF")
    ws["A1"].fill      = make_fill(C_TITLE_BG)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 32

    col_widths = [8, 20, 30, 35, 40, 35, 35, 12, 15, 12, 30]
    col_headers = [
        "TC ID", "Module", "Test Name", "Description",
        "Test Steps", "Expected Result", "Actual Result",
        "Status", "Timestamp", "Duration (s)", "Error / Notes"
    ]

    for ci, (hdr, wid) in enumerate(zip(col_headers, col_widths), start=1):
        cl = get_column_letter(ci)
        ws.column_dimensions[cl].width = wid
        ws[f"{cl}2"] = hdr
        ws[f"{cl}2"].font      = hdr_font(10)
        ws[f"{cl}2"].fill      = make_fill(C_HEADER_BG)
        ws[f"{cl}2"].alignment = center
        ws[f"{cl}2"].border    = border_header
    ws.row_dimensions[2].height = 28

    current_module = None
    row = 3

    for ri, r in enumerate(results):
        if r.module != current_module:
            current_module = r.module
            ws.merge_cells(f"A{row}:K{row}")
            ws[f"A{row}"] = f"  ▶  {current_module}"
            ws[f"A{row}"].font      = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
            ws[f"A{row}"].fill      = make_fill(C_SECTION_BG)
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 20
            row += 1

        status_color = C_PASS if r.status == "PASS" else (C_FAIL if r.status == "FAIL" else C_NOT_RUN)
        bg_fill = make_fill(C_ALT_ROW if ri % 2 == 0 else "FFFFFFFF")

        row_data = [
            r.tc_id, r.module, r.test_name, r.description,
            r.steps, r.expected, r.actual,
            r.status, r.timestamp, r.duration,
            r.error[:300] if r.error else ""
        ]

        for ci, val in enumerate(row_data, start=1):
            cl = get_column_letter(ci)
            ws[f"{cl}{row}"] = val
            ws[f"{cl}{row}"].border    = border_all
            ws[f"{cl}{row}"].alignment = left

            if ci == 8:
                ws[f"{cl}{row}"].font      = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
                ws[f"{cl}{row}"].fill      = make_fill(status_color)
                ws[f"{cl}{row}"].alignment = center
            elif ci == 1:
                ws[f"{cl}{row}"].font      = Font(name="Calibri", size=10, bold=True, color=C_ACCENT[2:])
                ws[f"{cl}{row}"].fill      = bg_fill
                ws[f"{cl}{row}"].alignment = center
            else:
                ws[f"{cl}{row}"].font = body_font()
                ws[f"{cl}{row}"].fill = bg_fill

        ws.row_dimensions[row].height = 45
        row += 1

    ws.freeze_panes = "A3"

    # --------------------------------------------------------------------------
    # SHEET 3 — FAILED TESTS
    # --------------------------------------------------------------------------
    ws_fail = wb.create_sheet("❌ Failed Tests")
    failed_tests = [r for r in results if r.status == "FAIL"]

    ws_fail.merge_cells("A1:K1")
    ws_fail["A1"] = f"❌  Failed Tests ({len(failed_tests)} of {total})"
    ws_fail["A1"].font      = Font(name="Calibri", size=13, bold=True, color="FFFFFFFF")
    ws_fail["A1"].fill      = make_fill("FF991B1B")
    ws_fail["A1"].alignment = center
    ws_fail.row_dimensions[1].height = 30

    for ci, (hdr, wid) in enumerate(zip(col_headers, col_widths), start=1):
        cl = get_column_letter(ci)
        ws_fail.column_dimensions[cl].width = wid
        ws_fail[f"{cl}2"] = hdr
        ws_fail[f"{cl}2"].font      = hdr_font(10)
        ws_fail[f"{cl}2"].fill      = make_fill(C_HEADER_BG)
        ws_fail[f"{cl}2"].alignment = center
        ws_fail[f"{cl}2"].border    = border_header
    ws_fail.row_dimensions[2].height = 28

    for ri, r in enumerate(failed_tests, start=3):
        row_data = [
            r.tc_id, r.module, r.test_name, r.description,
            r.steps, r.expected, r.actual,
            r.status, r.timestamp, r.duration,
            r.error[:300] if r.error else ""
        ]
        for ci, val in enumerate(row_data, start=1):
            cl = get_column_letter(ci)
            ws_fail[f"{cl}{ri}"] = val
            ws_fail[f"{cl}{ri}"].border    = border_all
            ws_fail[f"{cl}{ri}"].alignment = left
            ws_fail[f"{cl}{ri}"].font      = body_font()
            ws_fail[f"{cl}{ri}"].fill      = make_fill("FFFFF1F2" if ri % 2 == 0 else "FFFFFFFF")
            if ci == 8:
                ws_fail[f"{cl}{ri}"].fill      = make_fill(C_FAIL)
                ws_fail[f"{cl}{ri}"].font      = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
                ws_fail[f"{cl}{ri}"].alignment = center
        ws_fail.row_dimensions[ri].height = 55

    if not failed_tests:
        ws_fail.merge_cells("A3:K3")
        ws_fail["A3"] = "🎉  All tests passed! No failures."
        ws_fail["A3"].font      = Font(name="Calibri", size=14, bold=True, color="FF166534")
        ws_fail["A3"].fill      = make_fill("FFF0FDF4")
        ws_fail["A3"].alignment = center
        ws_fail.row_dimensions[3].height = 40

    wb.save(REPORT_PATH)
    print(f"\n[REPORT] Saved E2E Excel report at: {REPORT_PATH}")

# ==============================================================================
# Test Execution Workflow
# ==============================================================================
def execute_test_suite():
    global driver
    setup_driver()
    
    # Let the app boot
    print("[TEST_FLOW] Waiting for app to launch...")
    time.sleep(5)

    # --------------------------------------------------------------------------
    # Module 1 — Authentication Screen (Logged Out)
    # --------------------------------------------------------------------------
    
    # TC-04: Login with empty fields
    with test_case(
        "TC-04", "Authentication",
        "Login with empty fields",
        "Submit empty credentials, verify validation error is displayed.",
        "1. Find inputs\n2. Clear both fields\n3. Tap 'AUTHORIZE WORKSTATION SESSION'",
        "Validation error text 'Email and password cannot be blank.' is visible."
    ) as tr:
        fields = get_edit_texts()
        assert len(fields) >= 2, f"Expected 2 input fields, found {len(fields)}"
        fields[0].click()
        fields[0].clear()
        fields[1].click()
        fields[1].clear()
        hide_keyboard()
        click_by_text("AUTHORIZE WORKSTATION SESSION")
        err_msg = find_by_text("Email and password cannot be blank.")
        assert err_msg is not None, "Error message for blank fields not found"
        tr.actual = "Error message 'Email and password cannot be blank.' displayed correctly."

    # TC-02: Login with invalid email
    with test_case(
        "TC-02", "Authentication",
        "Login with invalid email",
        "Submit non-whitelisted email, verify access restrictions.",
        "1. Enter wrong@gmail.com\n2. Enter admin123\n3. Tap 'AUTHORIZE WORKSTATION SESSION'",
        "Error message 'Access restricted to saagar@gmail.com' is visible."
    ) as tr:
        fields = get_edit_texts()
        clear_and_type(fields[0], "wrong@gmail.com")
        clear_and_type(fields[1], "admin123")
        click_by_text("AUTHORIZE WORKSTATION SESSION")
        err_msg = find_by_text("Access restricted to saagar@gmail.com")
        assert err_msg is not None, "Error message for restricted email not found"
        tr.actual = "Restricted email correctly blocked with error message."

    # TC-03: Login with invalid password
    with test_case(
        "TC-03", "Authentication",
        "Login with invalid password",
        "Submit correct email and incorrect password, verify error message.",
        "1. Enter saagar@gmail.com\n2. Enter wrongpwd\n3. Tap 'AUTHORIZE WORKSTATION SESSION'",
        "Error message 'Incorrect password.' is visible."
    ) as tr:
        fields = get_edit_texts()
        clear_and_type(fields[0], "saagar@gmail.com")
        clear_and_type(fields[1], "wrongpwd")
        click_by_text("AUTHORIZE WORKSTATION SESSION")
        err_msg = find_by_text("Incorrect password.")
        assert err_msg is not None, "Error message for incorrect password not found"
        tr.actual = "Incorrect password blocked with error message."

    # TC-05: Navigate to Register screen
    with test_case(
        "TC-05", "Authentication",
        "Navigate to Register screen",
        "Verify register screen is reached by tapping navigation link.",
        "1. Tap 'Register Node' link",
        "Register Workstation screen is visible with text 'Register Workstation'."
    ) as tr:
        click_by_text("Register Node")
        title = find_by_text("Register Workstation")
        assert title is not None, "Register screen title not found"
        tr.actual = "Register screen reached successfully."

    # TC-07: Register with restricted email
    with test_case(
        "TC-07", "Authentication",
        "Register with restricted email",
        "Attempt registering non-whitelisted email, verify access restrictions.",
        "1. Fill Dr. Test, wrong@gmail.com, admin123\n2. Tap 'REGISTER SECURE STATION'",
        "Error message 'Access restricted to saagar@gmail.com' is visible."
    ) as tr:
        fields = get_edit_texts()
        assert len(fields) >= 3, f"Expected 3 input fields on Register screen, found {len(fields)}"
        clear_and_type(fields[0], "Dr. Test")
        clear_and_type(fields[1], "wrong@gmail.com")
        clear_and_type(fields[2], "admin123")
        click_by_text("REGISTER SECURE STATION")
        err_msg = find_by_text("Access restricted to saagar@gmail.com")
        assert err_msg is not None, "Error message for registration restricted email not found"
        tr.actual = "Register blocked for restricted email."

    # TC-06: Navigate back to Login from Register
    with test_case(
        "TC-06", "Authentication",
        "Navigate back to Login from Register",
        "Verify login screen can be returned to from registration screen.",
        "1. Tap 'Sign In' link",
        "Login screen title 'Clinical Workstation Portal' is visible."
    ) as tr:
        click_by_text("Sign In")
        title = find_by_text("Clinical Workstation Portal")
        assert title is not None, "Login screen title not found"
        tr.actual = "Returned to login screen successfully."

    # TC-01: Login with valid credentials
    with test_case(
        "TC-01", "Authentication",
        "Login with valid credentials",
        "Verify dashboard page loads after submitting correct credentials.",
        "1. Enter saagar@gmail.com\n2. Enter admin123\n3. Tap 'AUTHORIZE WORKSTATION SESSION'",
        "Dashboard screen is loaded and title 'Dashboard' is visible."
    ) as tr:
        fields = get_edit_texts()
        clear_and_type(fields[0], "saagar@gmail.com")
        clear_and_type(fields[1], "admin123")
        click_by_text("AUTHORIZE WORKSTATION SESSION")
        time.sleep(2)
        title = find_by_text("Dashboard", timeout=8)
        assert title is not None, "Dashboard screen failed to load"
        tr.actual = "Successfully logged in and reached Dashboard."

    # --------------------------------------------------------------------------
    # Module 2 — Dashboard Screen
    # --------------------------------------------------------------------------
    
    # TC-08: Dashboard loads after login
    with test_case(
        "TC-08", "Dashboard",
        "Dashboard loads after login",
        "Verify that dashboard renders UI elements post login.",
        "1. Check for 'Dashboard' screen title",
        "Dashboard title is visible."
    ) as tr:
        title = find_by_text("Dashboard")
        assert title is not None, "Dashboard title not found"
        tr.actual = "Dashboard screen verified."

    # TC-09: Stats cards are displayed
    with test_case(
        "TC-09", "Dashboard",
        "Stats cards are displayed",
        "Verify total sorted and inference speed cards are displayed.",
        "1. Look for 'TOTAL SORTED' and 'INFERENCE SPEED' texts",
        "Total Sorted and Inference Speed statistic cards are visible."
    ) as tr:
        total_sorted = find_by_text("TOTAL SORTED")
        inf_speed = find_by_text("INFERENCE SPEED")
        assert total_sorted is not None, "Total Sorted card not found"
        assert inf_speed is not None, "Inference Speed card not found"
        tr.actual = "Stats cards are visible."

    # TC-10: Modality cards displayed
    with test_case(
        "TC-10", "Dashboard",
        "Modality cards displayed",
        "Verify MRI, CT, X-Ray mod cards are present.",
        "1. Look for 'MRI SCANS', 'CT SCANS', 'X-RAYS', 'UNCLASSIFIED'",
        "All four modalities counts cards are displayed."
    ) as tr:
        mri = find_by_text("MRI SCANS")
        ct = find_by_text("CT SCANS")
        xray = find_by_text("X-RAYS")
        unknown = find_by_text("UNCLASSIFIED")
        assert mri is not None, "MRI card not found"
        assert ct is not None, "CT card not found"
        assert xray is not None, "X-Ray card not found"
        assert unknown is not None, "Unclassified card not found"
        tr.actual = "All modality stats cards are visible."

    # TC-11: Refresh button works
    with test_case(
        "TC-11", "Dashboard",
        "Refresh button works",
        "Verify clicking refresh updates stats and does not crash.",
        "1. Click refresh icon button",
        "Dashboard remains visible and content refreshes without errors."
    ) as tr:
        click_by_accessibility_id("Refresh stats")
        time.sleep(1)
        title = find_by_text("Dashboard")
        assert title is not None, "Dashboard title not found after refresh"
        tr.actual = "Refresh executed successfully."

    # --------------------------------------------------------------------------
    # Module 3 — Navigation Menu
    # --------------------------------------------------------------------------
    
    # TC-12: Navigate to AI Scan tab
    with test_case(
        "TC-12", "Navigation",
        "Navigate to AI Scan tab",
        "Verify bottom navigation bar navigates to AI Scan screen.",
        "1. Tap 'AI Scan' tab in bottom bar",
        "AI Classifier screen loads with title 'AI Classifier'."
    ) as tr:
        click_by_text("AI Scan")
        title = find_by_text("AI Classifier")
        assert title is not None, "AI Classifier title not found"
        tr.actual = "Successfully navigated to AI Scan tab."

    # TC-13: Navigate to Reports tab
    with test_case(
        "TC-13", "Navigation",
        "Navigate to Reports tab",
        "Verify bottom navigation bar navigates to Reports screen.",
        "1. Tap 'Reports' tab in bottom bar",
        "Reports screen loads with title 'Reports'."
    ) as tr:
        click_by_text("Reports")
        title = find_by_text("Reports")
        assert title is not None, "Reports title not found"
        tr.actual = "Successfully navigated to Reports tab."

    # TC-14: Navigate to Profile tab
    with test_case(
        "TC-14", "Navigation",
        "Navigate to Profile tab",
        "Verify bottom navigation bar navigates to Profile screen.",
        "1. Tap 'Profile' tab in bottom bar",
        "Profile screen loads with logout button 'CLOSE WORKSTATION SESSION'."
    ) as tr:
        click_by_text("Profile")
        btn = find_by_text("CLOSE WORKSTATION SESSION")
        assert btn is not None, "Logout button not found on Profile screen"
        tr.actual = "Successfully navigated to Profile tab."

    # TC-15: Navigate back to Dashboard
    with test_case(
        "TC-15", "Navigation",
        "Navigate back to Dashboard",
        "Verify bottom navigation bar navigates back to Dashboard screen.",
        "1. Tap 'Metrics' tab in bottom bar",
        "Dashboard screen loads with title 'Dashboard'."
    ) as tr:
        click_by_text("Metrics")
        title = find_by_text("Dashboard")
        assert title is not None, "Dashboard title not found"
        tr.actual = "Successfully returned to Dashboard tab."

    # --------------------------------------------------------------------------
    # Module 4 — Upload & AI Scan Screen
    # --------------------------------------------------------------------------
    
    # Navigate to Upload Screen
    click_by_text("AI Scan")
    time.sleep(1)

    # TC-16: Upload screen elements present
    with test_case(
        "TC-16", "Upload (AI Scan)",
        "Upload screen elements present",
        "Verify key upload screen components are displayed.",
        "1. Verify AI Classifier title\n2. Verify Select Medical Scan File text\n3. Verify DEMO PRESETS text",
        "All upload and preset elements are present."
    ) as tr:
        title = find_by_text("AI Classifier")
        select_text = find_by_text("Select Medical Scan File")
        preset_text = find_by_text("DEMO PRESETS (QUICK SCAN TEST)")
        assert title is not None, "AI Classifier title not found"
        assert select_text is not None, "Select Medical Scan File text not found"
        assert preset_text is not None, "DEMO PRESETS text not found"
        tr.actual = "AI Scan screen UI components verified."

    # TC-17: Demo preset selection
    with test_case(
        "TC-17", "Upload (AI Scan)",
        "Demo preset selection and analyze",
        "Verify preset scan selection and running AI neural network inference.",
        "1. Tap 'Chest X-Ray' preset\n2. Verify preview card\n3. Tap 'ANALYZE & CLASSIFY'\n4. Verify classification successful",
        "Preset selects, runs inference, and shows Analysis Successful modality results."
    ) as tr:
        click_by_text("Chest X-Ray")
        preview = find_by_text("Demo Preset Selected")
        assert preview is not None, "Preview screen for preset not shown"
        
        click_by_text("ANALYZE & CLASSIFY")
        # Wait for neural network simulated 2s latency
        time.sleep(3)
        
        success_msg = find_by_text("Analysis Successful", timeout=8)
        assert success_msg is not None, "Classification did not complete successfully"
        tr.actual = "Preset classified successfully with Analysis Successful result."

    # TC-18: Clear selected preset
    with test_case(
        "TC-18", "Upload (AI Scan)",
        "Clear selected preset",
        "Verify returning from results, selecting preset and clearing it works.",
        "1. Click 'SCAN ANOTHER IMAGE'\n2. Click 'Brain MRI' preset\n3. Click 'Clear'",
        "Preview screen closes and main selection options return."
    ) as tr:
        click_by_text("SCAN ANOTHER IMAGE")
        click_by_text("Brain MRI")
        preview = find_by_text("Demo Preset Selected")
        assert preview is not None, "Preview screen for Brain MRI not shown"
        
        click_by_text("Clear")
        time.sleep(1)
        select_text = find_by_text("Select Medical Scan File")
        assert select_text is not None, "Failed to clear preview back to main screen"
        tr.actual = "Preview cleared and main upload screen restored."

    # --------------------------------------------------------------------------
    # Module 5 — History (Reports) Screen
    # --------------------------------------------------------------------------
    
    # Navigate to Reports Screen
    click_by_text("Reports")
    time.sleep(1.5)

    # TC-19: History list renders
    with test_case(
        "TC-19", "History (Reports)",
        "History list renders",
        "Verify history list displays records successfully.",
        "1. Check if records found text exists\n2. Verify Reports header",
        "Historical reports list rendered with records."
    ) as tr:
        title = find_by_text("Reports")
        assert title is not None, "Reports title not found"
        # Search for text containing "records found"
        records = find_contains_text("records found")
        assert records is not None, "Records count summary not found"
        tr.actual = f"History screen loads. Found: '{records.text}'"

    # TC-20: Search bar is functional
    with test_case(
        "TC-20", "History (Reports)",
        "Search bar is functional",
        "Verify search bar accepts text input and filters list.",
        "1. Find search bar field\n2. Type 'Aarav'\n3. Verify records update",
        "Search field accepts input, filtering the list."
    ) as tr:
        search_field = driver.find_element(by=AppiumBy.CLASS_NAME, value="android.widget.EditText")
        clear_and_type(search_field, "Aarav")
        time.sleep(1)
        
        # Verify filtered records
        records_after = find_contains_text("records found")
        assert records_after is not None, "Records summary missing after search"
        
        # Clear search
        clear_and_type(search_field, "")
        time.sleep(1)
        tr.actual = f"Search filtered results successfully. Post-search status: '{records_after.text}'"

    # TC-21: Filter chips visible
    with test_case(
        "TC-21", "History (Reports)",
        "Filter chips visible and record dialog opens",
        "Verify filter chips are present and records open detail dialog overlay.",
        "1. Check chips 'All', 'MRI', 'CT Scan'\n2. Click first card item\n3. Verify detail dialog appears\n4. Tap 'CLOSE RECORD'",
        "Detail dialog overlay shows 'Clinical Scan Record' and closes correctly."
    ) as tr:
        # Check chips
        all_chip = find_by_text("All")
        mri_chip = find_by_text("MRI")
        ct_chip = find_by_text("CT Scan")
        assert all_chip is not None, "All filter chip not found"
        assert mri_chip is not None, "MRI filter chip not found"
        assert ct_chip is not None, "CT Scan filter chip not found"
        
        # Click MRI chip to filter
        mri_chip.click()
        time.sleep(1)
        
        # Click first item card
        cards = driver.find_elements(by=AppiumBy.XPATH, value="//*[contains(@text, 'PT-')]")
        assert len(cards) > 0, "No cards found under filtered list"
        cards[0].click()
        time.sleep(1)
        
        # Check details dialog
        dialog_title = find_by_text("Clinical Scan Record")
        assert dialog_title is not None, "Details dialog did not open"
        
        # Close dialog
        click_by_text("CLOSE RECORD")
        time.sleep(1)
        
        # Reset filter to All
        all_chip.click()
        time.sleep(1)
        tr.actual = "Filter chips and record details dialog work perfectly."

    # --------------------------------------------------------------------------
    # Module 6 — Profile Screen
    # --------------------------------------------------------------------------
    
    # Navigate to Profile Screen
    click_by_text("Profile")
    time.sleep(1)

    # TC-22: Profile shows user info
    with test_case(
        "TC-22", "Profile",
        "Profile shows user info",
        "Verify user profile displays login email saagar@gmail.com.",
        "1. Look for 'saagar@gmail.com' text in clinician profile details",
        "Dr. Saagar is displayed as clinician with email saagar@gmail.com."
    ) as tr:
        email_text = find_by_text("saagar@gmail.com")
        assert email_text is not None, "saagar@gmail.com email not displayed on Profile screen"
        tr.actual = "Doctor details and email saagar@gmail.com confirmed on profile."

    # TC-23: Edit profile dialog
    with test_case(
        "TC-23", "Profile",
        "Edit profile dialog",
        "Verify edit profile is restricted and displays locked notification.",
        "1. Tap 'EDIT PROFILE' button\n2. Verify 'Edit Credentials' dialog title\n3. Tap 'OK'",
        "Institution locked edit details dialog shows up and closes."
    ) as tr:
        click_by_text("EDIT PROFILE")
        dialog_title = find_by_text("Edit Credentials")
        assert dialog_title is not None, "Locked notification dialog not found"
        
        click_by_text("OK")
        time.sleep(1)
        tr.actual = "Institutional directory locking verified."

    # TC-24: Logout functionality
    with test_case(
        "TC-24", "Profile",
        "Logout functionality",
        "Verify CLOSE WORKSTATION SESSION exits to the workstation portal login.",
        "1. Tap 'CLOSE WORKSTATION SESSION'\n2. Verify transition back to Login screen",
        "Session closed and user returned to login screen."
    ) as tr:
        click_by_text("CLOSE WORKSTATION SESSION")
        time.sleep(2)
        login_title = find_by_text("Clinical Workstation Portal")
        assert login_title is not None, "Workstation portal login screen not found after logout"
        tr.actual = "Successfully logged out and session terminated."

    # --------------------------------------------------------------------------
    # Final Reporting
    # --------------------------------------------------------------------------
    generate_excel_report()
    
    # Terminate Appium session
    print("[TEST_FLOW] Terminating Appium WebDriver session...")
    driver.quit()

if __name__ == "__main__":
    try:
        execute_test_suite()
    except Exception as e:
        print(f"[ERROR] Suite execution failed: {e}")
        traceback.print_exc()
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
